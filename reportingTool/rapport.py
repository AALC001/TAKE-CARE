# Importer les librairies et fonctions nécessaires
from openpyxl import Workbook
import os
import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color


def rapport(data, libelle, filename):
    
    #style
    centre=Alignment(horizontal="center", vertical="center")
    font_title=Font(name='Times New Roman', size=22, bold=True)
    font_lib=Font(name='Arial Narrow', size=14, bold=True)
    font_tr=Font(name='Arial Narrow', size=13)
    font_li=Font(name='Calibri', size=10)
    barre_grise = PatternFill(start_color='404446',end_color='404446', fill_type='solid')
    
    # instanciation de la feuille dans la rame 
    wb = Workbook()
    ws = wb.active
    
    #formation des titres
    index=["B","AB","C","D","L","O","Q","R"]
    ws.merge_cells(F"{index[0]}3:{index[1]}3")
    ws.merge_cells(F"{index[0]}4:{index[1]}4")
    ws.merge_cells(F"{index[0]}10:{index[1]}10")
    ws[F"{index[0]}3"].font = font_title
    ws[F"{index[0]}3"]="RAPPORT MENSUEL D'ACTIVITES CONSEIL ET DEPISTAGE"
    ws[F"{index[0]}3"].alignment=centre
    ws[F"{index[0]}4"].fill=barre_grise
    ws[F"{index[0]}10"].fill=barre_grise
    for j in (5,6,7,8,9):
        #fusion
        ws.merge_cells(F"{index[0]}{j}:{index[2]}{j}")
        ws.merge_cells(F"{index[3]}{j}:{index[4]}{j}")
        ws.merge_cells(F"{index[5]}{j}:{index[6]}{j}")
        ws.merge_cells(F"{index[7]}{j}:{index[1]}{j}")
    #remplissage
    for j in (5,6,7,8,9):
        s=j-5
        ws[F"{index[0]}{j}"]=list(libelle.keys())[s]
        if s==0:
            ws[F"{index[3]}{j}"]=libelle[list(libelle.keys())[0]]
        else :
            ws[F"{index[3]}{j}"]=libelle[list(libelle.keys())[s]]
        ws[F"{index[0]}{j}"].font=font_lib  
        ws[F"{index[3]}{j}"].font=font_lib   
    for j in (5,6,7,8):
        ws[F"{index[5]}{j}"]=list(libelle.keys())[j]
        ws[F"{index[7]}{j}"]=libelle[list(libelle.keys())[j]]
        ws[F"{index[5]}{j}"].font=font_lib
        ws[F"{index[7]}{j}"].font=font_lib  

    # traçage du tableau pour chaque porte d'entreé donnée
    lignes=[
        "Nombre de clients conseillés pour le dépistage du VIH",
        "Nombre de clients dépistés pour le VIH",
        "Nombre de clients conseillés et dépistés pour le VIH ayant reçu le resultat",
        "Nombre de clients dépistés positif au VIH",
        "Nombre de clients dépistés positif au VIH ayant reçu le résultat du test",
        "Nombre de clients dépistés positif au VIH ayant bénéficié d'un CD4 ",
        "Nombre de clients dépistés VIH négatif"
        ]
    tranche=["<1","1-4","5-9","10-14","15-19",
            "20-24","25-29","30-34","35-39","40-44","45-49",
            "50+","<1","1-4","5-9","10-14","15-19",
            "20-24","25-29","30-34","35-39","40-44","45-49",
            "50+","Total"]
    
    #variable line per mini tableau
    line_table=11
    for i in data.keys():
        line_tranche=line_table+1
        #formation entete mini tableau
        for li in range(4,29):
            ws.cell(row = line_tranche, column = li).value=tranche[li-4]
            ws.cell(row = line_tranche, column = li).font=font_tr
            ws.cell(row = line_tranche, column = li).alignment=centre
        ent=["B","C","D","O","P","AA"] 
        for en,li in enumerate(range(line_table,line_table+9)):
            if en==0:
                ws.merge_cells(F"{ent[0]}{li}:{ent[1]}{li}")
                ws[F"{ent[0]}{li}"]="PORTE D'ENTREE"
                ws.merge_cells(F"{ent[2]}{li}:{ent[3]}{li}")
                ws[F"{ent[2]}{li}"]="Hommes"
                ws[F"{ent[2]}{li}"].alignment=centre
                ws.merge_cells(F"{ent[-2]}{li}:{ent[-1]}{li}")
                ws[F"{ent[-2]}{li}"]="Femmes"
                ws[F"{ent[-2]}{li}"].alignment=centre
            elif en==1:
                ws.merge_cells(F"{ent[0]}{li}:{ent[1]}{li}")
                ws[F"{ent[0]}{li}"]=i
            else:
                ws.merge_cells(F"{ent[0]}{li}:{ent[1]}{li}")
                ws[F"{ent[0]}{li}"]=lignes[en-2]
                ws[F"{ent[0]}{li}"].font=font_li
        index_gris=["B","AB"]
        ws.merge_cells(F"{index_gris[0]}{line_table+9}:{index_gris[1]}{line_table+9}")
        ws[F"{index_gris[0]}{line_table+9}"].fill=barre_grise
        line_table+=10

    #remplisssage de tous les tableaux
    line_remplis=11
    for i in data.keys():
        for en,li in enumerate(range(line_remplis+2,line_remplis+9)):
            minil=data[i][en]
            for k,j in enumerate(data[i][en]):
                ws.cell(row = li, column = k+4).value=j    
        line_remplis+=10
  
        
    filename =filename + '.xlsx'
    return wb.save(filename)
    #return filedialog.asksaveasfile(filename, wb)